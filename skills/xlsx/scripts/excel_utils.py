#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# +--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+ #
from __future__ import division
from __future__ import print_function
from __future__ import absolute_import
"""
Excel 工具类
提供常用的 Excel 操作方法
"""
# +--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+--+ #

import os
from typing import List, Dict, Any, Optional, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd


class ExcelUtils:
    """Excel 工具类"""

    # 预定义样式
    STYLES = {
        'title': {
            'font': Font(name='微软雅黑', size=16, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        },
        'header': {
            'font': Font(name='微软雅黑', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center')
        },
        'data': {
            'font': Font(name='微软雅黑', size=10),
            'alignment': Alignment(horizontal='left', vertical='center')
        },
        'number': {
            'font': Font(name='微软雅黑', size=10),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': '#,##0.00'
        },
        'currency': {
            'font': Font(name='微软雅黑', size=10),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': '"¥"#,##0.00'
        },
        'percent': {
            'font': Font(name='微软雅黑', size=10),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'number_format': '0.00%'
        }
    }

    THIN_BORDER = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    @staticmethod
    def create_workbook() -> Workbook:
        """
        创建新的工作簿

        Returns:
            Workbook: 新的工作簿对象
        """
        wb = Workbook()
        # 删除默认工作表
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb

    @staticmethod
    def load_workbook(filepath: str, read_only: bool = False) -> Workbook:
        """
        加载现有工作簿

        Args:
            filepath: 文件路径
            read_only: 是否以只读模式加载

        Returns:
            Workbook: 工作簿对象
        """
        return load_workbook(filepath, read_only=read_only)

    @staticmethod
    def apply_style(cell, style_name: str, border: bool = False):
        """
        应用样式到单元格

        Args:
            cell: 单元格对象
            style_name: 样式名称（从 STYLES 中选择）
            border: 是否添加边框
        """
        if style_name not in ExcelUtils.STYLES:
            return

        style = ExcelUtils.STYLES[style_name]

        if 'font' in style:
            cell.font = style['font']
        if 'fill' in style:
            cell.fill = style['fill']
        if 'alignment' in style:
            cell.alignment = style['alignment']
        if 'number_format' in style:
            cell.number_format = style['number_format']

        if border:
            cell.border = ExcelUtils.THIN_BORDER

    @staticmethod
    def set_cell_value(ws: Worksheet, row: int, col: int, value: Any,
                      style_name: Optional[str] = None, border: bool = False):
        """
        设置单元格值和样式

        Args:
            ws: 工作表对象
            row: 行号（从1开始）
            col: 列号（从1开始）
            value: 单元格值
            style_name: 样式名称
            border: 是否添加边框
        """
        cell = ws.cell(row=row, column=col, value=value)
        if style_name:
            ExcelUtils.apply_style(cell, style_name, border)

    @staticmethod
    def add_header(ws: Worksheet, headers: List[str], row: int = 1,
                  start_col: int = 1, border: bool = True):
        """
        添加表头

        Args:
            ws: 工作表对象
            headers: 表头列表
            row: 起始行号
            start_col: 起始列号
            border: 是否添加边框
        """
        for col_idx, header in enumerate(headers, start=start_col):
            ExcelUtils.set_cell_value(ws, row, col_idx, header, 'header', border)

    @staticmethod
    def add_data(ws: Worksheet, data: List[List[Any]], start_row: int = 2,
                 start_col: int = 1, style_name: str = 'data', border: bool = True):
        """
        添加数据行

        Args:
            ws: 工作表对象
            data: 二维数据列表
            start_row: 起始行号
            start_col: 起始列号
            style_name: 样式名称
            border: 是否添加边框

        Returns:
            int: 最后写入的行号
        """
        for row_idx, row_data in enumerate(data, start=start_row):
            for col_idx, value in enumerate(row_data, start=start_col):
                ExcelUtils.set_cell_value(ws, row_idx, col_idx, value, style_name, border)
        return start_row + len(data) - 1

    @staticmethod
    def merge_cells(ws: Worksheet, start_row: int, start_col: int,
                    end_row: int, end_col: int, value: Any = None,
                    style_name: Optional[str] = None):
        """
        合并单元格

        Args:
            ws: 工作表对象
            start_row: 起始行
            start_col: 起始列
            end_row: 结束行
            end_col: 结束列
            value: 合并后的值
            style_name: 样式名称
        """
        ws.merge_cells(start_row=start_row, start_column=start_col,
                       end_row=end_row, end_column=end_col)

        if value:
            cell = ws.cell(row=start_row, column=start_col)
            cell.value = value
            if style_name:
                ExcelUtils.apply_style(cell, style_name)

    @staticmethod
    def auto_adjust_column_width(ws: Worksheet, min_width: int = 8, max_width: int = 50):
        """
        自动调整列宽

        Args:
            ws: 工作表对象
            min_width: 最小列宽
            max_width: 最大列宽
        """
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value is not None else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass

            adjusted_width = max(min(min_width, max(max_length + 2, min_width)), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def freeze_panes(ws: Worksheet, cell: str = 'A2'):
        """
        冻结窗格

        Args:
            ws: 工作表对象
            cell: 冻结位置单元格（例如 'A2' 冻结首行）
        """
        ws.freeze_panes = cell

    @staticmethod
    def enable_auto_filter(ws: Worksheet, start_row: int = 1,
                          start_col: int = 1, end_row: Optional[int] = None,
                          end_col: Optional[int] = None):
        """
        启用自动筛选

        Args:
            ws: 工作表对象
            start_row: 起始行
            start_col: 起始列
            end_row: 结束行（None 表示使用最大行）
            end_col: 结束列（None 表示使用最大列）
        """
        if end_row is None:
            end_row = ws.max_row
        if end_col is None:
            end_col = ws.max_column

        start_cell = get_column_letter(start_col) + str(start_row)
        end_cell = get_column_letter(end_col) + str(end_row)
        ws.auto_filter.ref = f"{start_cell}:{end_cell}"

    @staticmethod
    def save_workbook(wb: Workbook, filepath: str, overwrite: bool = False):
        """
        保存工作簿

        Args:
            wb: 工作簿对象
            filepath: 文件路径
            overwrite: 是否覆盖已存在的文件
        """
        if os.path.exists(filepath) and not overwrite:
            raise FileExistsError(f"文件已存在: {filepath}")

        # 确保目录存在
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        wb.save(filepath)

    @staticmethod
    def read_csv_to_dataframe(filepath: str, encoding: str = 'utf-8',
                              **kwargs) -> pd.DataFrame:
        """
        读取 CSV 文件到 DataFrame

        Args:
            filepath: 文件路径
            encoding: 文件编码
            **kwargs: pandas.read_csv 的其他参数

        Returns:
            DataFrame: 数据框
        """
        return pd.read_csv(filepath, encoding=encoding, **kwargs)

    @staticmethod
    def write_dataframe_to_excel(df: pd.DataFrame, filepath: str,
                                  sheet_name: str = 'Sheet1', index: bool = False,
                                  **kwargs):
        """
        将 DataFrame 写入 Excel

        Args:
            df: 数据框
            filepath: 文件路径
            sheet_name: 工作表名称
            index: 是否写入索引
            **kwargs: DataFrame.to_excel 的其他参数
        """
        # 确保目录存在
        os.makedirs(os.path.dirname(filepath), exist_ok=True)

        df.to_excel(filepath, sheet_name=sheet_name, index=index, **kwargs)

    @staticmethod
    def read_excel_to_dataframe(filepath: str, sheet_name: Optional[str] = None,
                                 **kwargs) -> Union[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """
        读取 Excel 文件到 DataFrame

        Args:
            filepath: 文件路径
            sheet_name: 工作表名称（None 表示读取所有工作表）
            **kwargs: pandas.read_excel 的其他参数

        Returns:
            DataFrame 或 DataFrame 字典
        """
        if sheet_name:
            return pd.read_excel(filepath, sheet_name=sheet_name, **kwargs)
        else:
            return pd.read_excel(filepath, sheet_name=None, **kwargs)

    @staticmethod
    def create_quotation_table(filepath: str, title: str, headers: List[str],
                               data: List[List[Any]], auto_width: bool = True):
        """
        创建报价表

        Args:
            filepath: 文件路径
            title: 表格标题
            headers: 表头列表
            data: 数据列表
            auto_width: 是否自动调整列宽
        """
        # 创建工作簿
        wb = ExcelUtils.create_workbook()
        ws = wb.active
        ws.title = "报价表"

        # 添加标题
        ExcelUtils.merge_cells(ws, 1, 1, 1, len(headers), title, 'title')

        # 添加表头
        ExcelUtils.add_header(ws, headers, row=3, border=True)

        # 添加数据
        ExcelUtils.add_data(ws, data, start_row=4, start_col=1, style_name='data', border=True)

        # 自动调整列宽
        if auto_width:
            ExcelUtils.auto_adjust_column_width(ws)

        # 冻结窗格
        ExcelUtils.freeze_panes(ws, 'A4')

        # 启用筛选
        ExcelUtils.enable_auto_filter(ws, start_row=3)

        # 保存文件
        ExcelUtils.save_workbook(wb, filepath, overwrite=True)

        print(f"报价表创建成功: {filepath}")


def main():
    """测试函数"""
    # 测试创建报价表
    title = "测试报价表"
    headers = ['项目', '描述', '价格(元)', '单位']
    data = [
        ['项目1', '这是项目1的描述', '1000', '个'],
        ['项目2', '这是项目2的描述', '2000', '套'],
        ['项目3', '这是项目3的描述', '3000', '组']
    ]

    filepath = os.path.join(os.path.dirname(__file__), '..', 'test_output.xlsx')
    ExcelUtils.create_quotation_table(filepath, title, headers, data)


if __name__ == '__main__':
    main()
